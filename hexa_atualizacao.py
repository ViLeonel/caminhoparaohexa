"""Orquestração da Central de Atualização.

A camada aceita arquivos estruturados, valida o conteúdo e persiste somente
documentos sazonais ou calendários. O cadastro canônico dos jogadores não é
alterado por este módulo.
"""

from __future__ import annotations

import csv
import io
import json
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Mapping, Sequence

from hexa_calendarios import ResultadoCalendario, atualizar_calendario
from hexa_estatisticas import ResultadoTemporada, atualizar_temporada

__all__ = [
    "ARQUIVOS_ACEITOS",
    "AtualizacaoFileError",
    "DocumentoRepository",
    "PreviaAtualizacao",
    "carregar_registros_arquivo",
    "preparar_calendario",
    "preparar_temporada",
]

ARQUIVOS_ACEITOS: tuple[str, ...] = ("json", "csv", "xlsx")


class AtualizacaoFileError(ValueError):
    """Indica arquivo de atualização ausente, inválido ou inseguro."""


@dataclass(frozen=True, slots=True)
class PreviaAtualizacao:
    tipo: str
    ano: int
    documento: dict[str, Any]
    recebidos: int
    atualizados: int
    avisos: tuple[str, ...] = ()


class DocumentoRepository:
    """Persistência JSON com backup, validação e escrita atômica."""

    def __init__(self, diretorio: Path, prefixo: str) -> None:
        self.diretorio = Path(diretorio)
        self.prefixo = str(prefixo).strip()
        if not self.prefixo:
            raise ValueError("prefixo é obrigatório.")

    def caminho(self, ano: int) -> Path:
        return self.diretorio / f"{self.prefixo}_{int(ano)}.json"

    def carregar(self, ano: int) -> dict[str, Any] | None:
        caminho = self.caminho(ano)
        if not caminho.exists():
            return None
        try:
            documento = json.loads(caminho.read_text(encoding="utf-8-sig"))
        except (OSError, json.JSONDecodeError) as erro:
            raise AtualizacaoFileError(
                f"{caminho.name} está inválido e não foi sobrescrito."
            ) from erro
        if not isinstance(documento, dict):
            raise AtualizacaoFileError(
                f"{caminho.name} precisa conter um objeto JSON."
            )
        return documento

    def salvar(self, ano: int, documento: Mapping[str, Any]) -> Path:
        self.diretorio.mkdir(parents=True, exist_ok=True)
        destino = self.caminho(ano)
        temporario = destino.with_suffix(".json.tmp")
        backup = destino.with_suffix(".json.bak")
        conteudo = json.dumps(
            documento,
            ensure_ascii=False,
            indent=2,
            sort_keys=False,
        )
        temporario.write_text(conteudo + "\n", encoding="utf-8")
        json.loads(temporario.read_text(encoding="utf-8"))
        if destino.exists():
            shutil.copy2(destino, backup)
        temporario.replace(destino)
        return destino


def _json_registros(conteudo: bytes) -> list[dict[str, Any]]:
    try:
        dados = json.loads(conteudo.decode("utf-8-sig"))
    except (UnicodeDecodeError, json.JSONDecodeError) as erro:
        raise AtualizacaoFileError("O arquivo JSON é inválido.") from erro
    if isinstance(dados, list):
        registros = dados
    elif isinstance(dados, dict):
        registros = (
            dados.get("registros")
            or dados.get("estatisticas")
            or dados.get("jogos")
            or dados.get("dados")
        )
    else:
        registros = None
    if not isinstance(registros, list):
        raise AtualizacaoFileError(
            "O JSON precisa conter uma lista ou uma chave registros/estatisticas/jogos."
        )
    if not all(isinstance(item, Mapping) for item in registros):
        raise AtualizacaoFileError("Todos os registros precisam ser objetos.")
    return [dict(item) for item in registros]


def _csv_registros(conteudo: bytes) -> list[dict[str, Any]]:
    try:
        texto = conteudo.decode("utf-8-sig")
    except UnicodeDecodeError as erro:
        raise AtualizacaoFileError("O CSV precisa usar UTF-8.") from erro
    amostra = texto[:4096]
    try:
        dialeto = csv.Sniffer().sniff(amostra, delimiters=",;\t")
    except csv.Error:
        dialeto = csv.excel
    leitor = csv.DictReader(io.StringIO(texto), dialect=dialeto)
    if not leitor.fieldnames:
        raise AtualizacaoFileError("O CSV não possui cabeçalho.")
    return [dict(linha) for linha in leitor]


def _xlsx_registros(conteudo: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as erro:
        raise AtualizacaoFileError(
            "A leitura XLSX requer a dependência openpyxl."
        ) from erro

    try:
        planilha = load_workbook(
            io.BytesIO(conteudo),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as erro:
        raise AtualizacaoFileError("A planilha XLSX não pôde ser lida.") from erro
    try:
        aba = planilha.active
        linhas = aba.iter_rows(values_only=True)
        cabecalho_bruto = next(linhas, None)
        if not cabecalho_bruto:
            raise AtualizacaoFileError("A planilha está vazia.")
        cabecalho = [str(valor or "").strip() for valor in cabecalho_bruto]
        if not all(cabecalho):
            raise AtualizacaoFileError(
                "Todas as colunas do cabeçalho precisam ter nome."
            )
        registros = [
            {
                cabecalho[indice]: valor
                for indice, valor in enumerate(linha)
                if indice < len(cabecalho)
            }
            for linha in linhas
            if any(valor not in (None, "") for valor in linha)
        ]
    finally:
        planilha.close()
    return registros


def carregar_registros_arquivo(
    *,
    nome_arquivo: str,
    conteudo: bytes,
    tamanho_maximo: int = 5 * 1024 * 1024,
) -> list[dict[str, Any]]:
    """Lê JSON, CSV ou XLSX sem executar fórmulas, macros ou código."""
    if not conteudo:
        raise AtualizacaoFileError("O arquivo está vazio.")
    if len(conteudo) > tamanho_maximo:
        raise AtualizacaoFileError("O arquivo excede o limite de 5 MB.")
    extensao = Path(nome_arquivo).suffix.lower().lstrip(".")
    if extensao not in ARQUIVOS_ACEITOS:
        raise AtualizacaoFileError(
            "Formato não permitido. Use JSON, CSV ou XLSX."
        )
    if extensao == "json":
        return _json_registros(conteudo)
    if extensao == "csv":
        return _csv_registros(conteudo)
    return _xlsx_registros(conteudo)


def preparar_temporada(
    *,
    ano: int,
    registros: Sequence[Mapping[str, Any]],
    jogadores: Mapping[str, Mapping[str, Any]],
    documento_anterior: Mapping[str, Any] | None,
    fonte: str,
    atualizado_por: str,
    agora: datetime | None = None,
) -> PreviaAtualizacao:
    resultado: ResultadoTemporada = atualizar_temporada(
        temporada=ano,
        registros=registros,
        jogadores=jogadores,
        documento_anterior=documento_anterior,
        fonte=fonte,
        atualizado_por=atualizado_por,
        agora=agora,
    )
    avisos = (
        *(
            f"Atleta não encontrado: {nome}"
            for nome in resultado.nao_encontrados
        ),
        *(
            f"Registro duplicado descartado: {item}"
            for item in resultado.duplicados
        ),
    )
    return PreviaAtualizacao(
        tipo="temporada",
        ano=int(ano),
        documento=resultado.documento,
        recebidos=resultado.recebidos,
        atualizados=resultado.atualizados,
        avisos=tuple(avisos),
    )


def preparar_calendario(
    *,
    ano: int,
    registros: Sequence[Mapping[str, Any]],
    documento_anterior: Mapping[str, Any] | None,
    fonte: str,
    atualizado_por: str,
    agora: datetime | None = None,
) -> PreviaAtualizacao:
    resultado: ResultadoCalendario = atualizar_calendario(
        ano=ano,
        jogos=registros,
        documento_anterior=documento_anterior,
        fonte=fonte,
        atualizado_por=atualizado_por,
        agora=agora,
    )
    avisos = tuple(
        f"Jogo duplicado descartado: {item}"
        for item in resultado.duplicados
    )
    return PreviaAtualizacao(
        tipo="calendario",
        ano=int(ano),
        documento=resultado.documento,
        recebidos=resultado.recebidos,
        atualizados=resultado.atualizados,
        avisos=avisos,
    )
